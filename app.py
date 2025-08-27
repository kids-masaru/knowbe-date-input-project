# app.py (段階的処理対応版)

import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from google.oauth2.service_account import Credentials
from datetime import datetime
import time
import re

# --- ページ設定 ---
st.set_page_config(
    page_title="Excel直接更新システム",
    page_icon="📎",
    layout="centered"
)

# --- CSS ---
st.markdown("""
<style>
body { font-family: 'Noto Sans JP', sans-serif; }
.main .block-container { padding-top: 2rem; }
h1 { border-bottom: 2px solid #2563eb; padding-bottom: 0.5rem; }
.step-info {
    background-color: #f0f8ff;
    padding: 1rem;
    border-radius: 0.5rem;
    border-left: 4px solid #2563eb;
    margin: 1rem 0;
}
</style>
""", unsafe_allow_html=True)

st.title("📎 Excel直接更新システム（段階処理対応）")
st.markdown("更新元ファイル（CSV/Excel）をアップロードすると、Google Drive上の指定のExcelファイルを段階的に更新します。")

# --- ヘルパー関数 ---
def check_secrets():
    """必要なsecrets設定が存在するかチェック"""
    missing_keys = []
    if "gcp_service_account" not in st.secrets:
        missing_keys.append("gcp_service_account")
    return missing_keys

def get_google_creds():
    try:
        creds = Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        return creds
    except Exception as e:
        st.error(f"Googleへの認証に失敗しました: {e}")
        return None

def extract_file_id_from_url(url_or_id):
    """URLまたはファイルIDからファイルIDを抽出"""
    if not url_or_id:
        return ""
    if len(url_or_id) > 10 and '/' not in url_or_id:
        return url_or_id.strip()
    
    patterns = [
        r'/file/d/([a-zA-Z0-9-_]+)',
        r'id=([a-zA-Z0-9-_]+)',
        r'/folders/([a-zA-Z0-9-_]+)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, url_or_id)
        if match:
            return match.group(1)
    
    return url_or_id.strip()

def col_num_to_letter(col_num):
    """列番号を文字に変換 (1=A, 26=Z, 27=AA)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + col_num % 26) + result
        col_num //= 26
    return result

# --- 設定チェック ---
missing_keys = check_secrets()
if missing_keys:
    st.error(f"""
    **設定エラー:** 以下の設定が不足しています：
    - {', '.join(missing_keys)}
    """)
    st.stop()

# --- Google Drive ファイルID の設定 ---
st.subheader("📁 更新対象のGoogle DriveファイルID")

default_file_id = ""
try:
    default_file_id = st.secrets.get("target_excel_file_id", "")
except:
    pass

col1, col2 = st.columns([3, 1])
with col1:
    file_id = st.text_input(
        "Google DriveファイルのIDまたはURLを入力してください",
        value=default_file_id,
        placeholder="例: 1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms",
        help="DriveのURL: https://drive.google.com/file/d/【この部分がID】/view"
    )

file_id = extract_file_id_from_url(file_id)

if not file_id:
    st.warning("**ファイルIDが入力されていません。**")
    st.info("""
    📝 **ファイルIDの取得方法:**
    1. Google Driveで対象のExcelファイルを開く
    2. URLをコピー: `https://drive.google.com/file/d/【この部分】/view`
    3. 上記の【この部分】がファイルIDです
    """)
else:
    st.success(f"**対象ファイルID:** `{file_id}`")

# --- 処理モード選択 ---
st.subheader("🔄 処理モード選択")

process_mode = st.radio(
    "処理方法を選択してください：",
    options=["一括処理（1枚目のみ更新）", "段階処理（2枚目→3枚目のコピーも実行）"],
    help="段階処理は1枚目更新後、Excel関数の計算を待ってから2枚目→3枚目のコピーを実行します"
)

# 段階処理の設定
if process_mode == "段階処理（2枚目→3枚目のコピーも実行）":
    st.markdown('<div class="step-info">', unsafe_allow_html=True)
    st.markdown("""
    **📋 段階処理の流れ:**
    1. **1枚目更新** - アップロードしたデータを1枚目に貼り付け
    2. **中間保存** - Driveに保存してExcel関数を計算させる
    3. **待機時間** - 関数計算の完了を待つ
    4. **再取得** - 計算済みのファイルをダウンロード
    5. **コピー処理** - 2枚目の計算結果を3枚目にコピー
    6. **最終保存** - 完了したファイルを保存
    """)
    st.markdown('</div>', unsafe_allow_html=True)
    
    wait_time = st.slider(
        "計算待機時間（秒）", 
        min_value=1, 
        max_value=15, 
        value=5, 
        help="1枚目更新後、Excel関数の計算を待つ時間"
    )
    
    st.info(f"⏱️ 設定された待機時間: **{wait_time}秒**")

# --- ファイルアップロード ---
st.subheader("📁 ファイルアップロード")
uploaded_file = st.file_uploader(
    "更新元となるファイル（CSVまたはExcel）を選択してください",
    type=['csv', 'xlsx', 'xls'],
    label_visibility="collapsed"
)

if uploaded_file:
    st.info(f"**選択中のファイル:** {uploaded_file.name}")

# --- 実行ボタン ---
is_pressed = st.button(
    "🚀 Drive上のExcelを更新実行", 
    type="primary", 
    use_container_width=True, 
    disabled=(uploaded_file is None or not file_id)
)

# 処理状況表示用
if 'processing_log' not in st.session_state:
    st.session_state.processing_log = []

# --- メイン処理 ---
if is_pressed:
    if uploaded_file is None:
        st.error("エラー: ファイルがアップロードされていません。")
        st.stop()
    
    if not file_id:
        st.error("エラー: Google DriveのファイルIDが入力されていません。")
        st.stop()
    
    # 処理ログをリセット
    st.session_state.processing_log = []
    
    start_time = time.time()
    creds = get_google_creds()

    if creds:
        try:
            # プログレスバー用のコンテナ
            progress_container = st.container()
            log_container = st.container()
            
            with progress_container:
                if process_mode == "一括処理（1枚目のみ更新）":
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
            
            # Drive APIサービスを構築
            drive_service = build('drive', 'v3', credentials=creds)
            
            # アップロードファイルの読み込み
            status_text.text("📄 アップロードファイルを読み込み中...")
            file_extension = uploaded_file.name.lower()
            if file_extension.endswith('.csv'):
                source_df = pd.read_csv(uploaded_file)
            elif file_extension.endswith(('.xlsx', '.xls')):
                source_df = pd.read_excel(uploaded_file, sheet_name=0)
            else:
                st.error(f"サポートされていないファイル形式です: {uploaded_file.name}")
                st.stop()
            
            progress_bar.progress(0.1)
            
            # Drive上のファイルをダウンロード
            status_text.text("☁️ Drive上のファイルをダウンロード中...")
            request = drive_service.files().get_media(fileId=file_id)
            file_content_bytes = request.execute()
            fh = io.BytesIO(file_content_bytes)
            
            progress_bar.progress(0.2)
            
            # ワークブックを読み込み
            status_text.text("📊 Excelワークブックを読み込み中...")
            workbook = openpyxl.load_workbook(fh, keep_vba=True)
            
            progress_bar.progress(0.3)
            
            # 1枚目のシートを更新
            status_text.text("✏️ 1枚目のシートを更新中...")
            sheet_to_update = workbook.worksheets[0]
            
            # 既存データをクリア（ヘッダーは保持）
            if sheet_to_update.max_row > 1:
                sheet_to_update.delete_rows(2, sheet_to_update.max_row)
            
            # 新しいデータを書き込み
            start_row = 2 if sheet_to_update.max_row >= 1 else 1
            for r_idx, row in enumerate(dataframe_to_rows(source_df, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    sheet_to_update.cell(row=r_idx, column=c_idx, value=value)
            
            progress_bar.progress(0.5)
            
            if process_mode == "一括処理（1枚目のみ更新）":
                # 一括処理の場合はそのまま保存
                status_text.text("💾 ファイルを保存中...")
                output_buffer = io.BytesIO()
                workbook.save(output_buffer)
                output_buffer.seek(0)
                
                media = MediaIoBaseUpload(output_buffer, mimetype='application/vnd.ms-excel.sheet.macroEnabled.12')
                drive_service.files().update(fileId=file_id, media_body=media).execute()
                
                progress_bar.progress(1.0)
                status_text.text("✅ 処理完了！")
                
            else:
                # 段階処理の場合
                status_text.text("💾 中間保存中...")
                output_buffer = io.BytesIO()
                workbook.save(output_buffer)
                output_buffer.seek(0)
                
                media = MediaIoBaseUpload(output_buffer, mimetype='application/vnd.ms-excel.sheet.macroEnabled.12')
                drive_service.files().update(fileId=file_id, media_body=media).execute()
                
                progress_bar.progress(0.6)
                
                # 待機時間
                status_text.text(f"⏳ Excel関数の計算待機中... ({wait_time}秒)")
                wait_progress = st.progress(0)
                for i in range(wait_time):
                    time.sleep(1)
                    wait_progress.progress((i + 1) / wait_time)
                    status_text.text(f"⏳ Excel関数の計算待機中... ({wait_time - i - 1}秒)")
                
                progress_bar.progress(0.7)
                
                # 計算済みファイルを再取得
                status_text.text("🔄 計算済みファイルを再取得中...")
                request = drive_service.files().get_media(fileId=file_id)
                updated_file_content = request.execute()
                updated_fh = io.BytesIO(updated_file_content)
                
                # 計算済みワークブックを読み込み
                calculated_workbook = openpyxl.load_workbook(updated_fh, keep_vba=True, data_only=True)
                final_workbook = openpyxl.load_workbook(updated_fh, keep_vba=True)
                
                progress_bar.progress(0.8)
                
                # 2枚目→3枚目のコピー処理
                status_text.text("📋 2枚目→3枚目のコピー処理中...")
                if len(calculated_workbook.worksheets) >= 3:
                    sheet2_calculated = calculated_workbook.worksheets[1]  # 計算済みの2枚目
                    sheet3_write = final_workbook.worksheets[2]            # 書き込み用3枚目
                    
                    # 2枚目の名前リスト（奇数行のみ: 7, 9, 11...）
                    names_sheet2 = {}
                    for row in range(7, min(sheet2_calculated.max_row + 1, 100), 2):
                        name = sheet2_calculated.cell(row=row, column=2).value  # B列
                        if name and str(name).strip():
                            clean_name = str(name).strip()
                            names_sheet2[clean_name] = row
                    
                    # 3枚目の名前リスト（N列、19行目以降）
                    names_sheet3 = {}
                    for row in range(19, min(sheet3_write.max_row + 1, 200)):
                        name = sheet3_write.cell(row=row, column=14).value  # N列
                        if name and str(name).strip():
                            clean_name = str(name).strip()
                            names_sheet3[clean_name] = row
                    
                    # マッチした名前のコピー処理
                    copy_count = 0
                    copy_log = []
                    
                    for name, sheet2_row in names_sheet2.items():
                        if name in names_sheet3:
                            sheet3_row = names_sheet3[name]
                            copy_log.append(f"名前マッチ: {name} (2枚目{sheet2_row}行 → 3枚目{sheet3_row}行)")
                            
                            # C列以降のデータをコピー（計算済みの値）
                            # 2枚目のC列（3列目）→ 3枚目のO列（15列目）以降に対応
                            for col in range(3, min(sheet2_calculated.max_column + 1, 95)):
                                calculated_value = sheet2_calculated.cell(row=sheet2_row, column=col).value
                                
                                # 正しい列位置計算: 2枚目のC列→3枚目のO列（15列目）
                                target_col = col + 12  # C(3)→O(15), D(4)→P(16), E(5)→Q(17)...
                                
                                # 3枚目の列範囲を確認（O列=15列目以降）
                                if target_col >= 15 and target_col <= 200:  # O列以降かつ妥当な範囲
                                    sheet3_write.cell(row=sheet3_row, column=target_col).value = calculated_value
                                    
                                    if calculated_value is not None:
                                        copy_count += 1
                                        # 詳細ログ（最初の5個のみ）
                                        if len(copy_log) < 20:
                                            source_col_letter = col_num_to_letter(col)
                                            target_col_letter = col_num_to_letter(target_col)
                                            copy_log.append(f"    {source_col_letter}{sheet2_row}({calculated_value})→{target_col_letter}{sheet3_row}")
                    
                    # コピー結果をログに追加
                    with log_container:
                        st.success(f"✅ {copy_count}個の計算済みセルを2枚目から3枚目にコピーしました")
                        st.info(f"📊 マッチした名前: {len(names_sheet2)} → {len(names_sheet3)} 中 {len(set(names_sheet2.keys()) & set(names_sheet3.keys()))} 件")
                        
                        if copy_log:
                            with st.expander("📋 コピー詳細ログ"):
                                for log in copy_log[:20]:  # 最初の20件
                                    st.text(log)
                
                progress_bar.progress(0.9)
                
                # 最終保存
                status_text.text("💾 最終保存中...")
                final_buffer = io.BytesIO()
                final_workbook.save(final_buffer)
                final_buffer.seek(0)
                
                final_media = MediaIoBaseUpload(final_buffer, mimetype='application/vnd.ms-excel.sheet.macroEnabled.12')
                drive_service.files().update(fileId=file_id, media_body=final_media).execute()
                
                progress_bar.progress(1.0)
                status_text.text("✅ 段階処理完了！")
            
            # 完了メッセージ
            end_time = time.time()
            processing_time = end_time - start_time
            now_str = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
            
            st.success(f"""
            **🎉 更新完了！**
            
            - **処理モード**: {process_mode}
            - **完了日時**: {now_str}
            - **処理時間**: {processing_time:.1f}秒
            - **対象ファイル**: `{file_id}`
            """)

        except Exception as e:
            st.error(f"**エラーが発生しました:** {e}")
            with st.expander("詳細なエラー情報"):
                import traceback
                st.text(traceback.format_exc())
